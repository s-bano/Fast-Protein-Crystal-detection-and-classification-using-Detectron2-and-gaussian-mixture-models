import h5py, time, joblib, csv, os
import numpy as np
from collections import Counter
from sklearn.decomposition import PCA
from sklearn.preprocessing import StandardScaler
from sklearn.mixture import GaussianMixture
from sklearn.pipeline import Pipeline
from pathlib import Path
from openpyxl import Workbook
from collections import defaultdict


STANDARD_MODEL = "model_classif_0812.pkl"
STANDARD_SCALER = StandardScaler()
PCA_DIM = 64



# This function extracts the features from a h5 file
def extract_h5(h5_path):
    """
    This function extracts the features from a h5 file
    
    outputs:
        - a list of features (non concatenated and non-scaled) ready to be classified (list of arrays),
        - a list of corresponding image names.
    """
    all_box_features = []
    image_names = []
    with h5py.File(h5_path, "r") as f:
        for img_name in f.keys():
            box_feats = f[img_name]["box_features"][:]  # shape (N, D)
            all_box_features.append(box_feats)
            image_names.append(img_name)
            
    
    return all_box_features, image_names


# Recapitulatif de la proportion de cristaux classifies dans chaque cluster
def proportions(lst):
    count = Counter(lst)
    total = len(lst)
    return {
        "Cluster id " + str(int(k)): round((v / total) * 100, 2)
        for k, v in count.items()
    }


# Conactene les box_features, normalize les donnees avec un scaler et applique une reudtcion PCA
def _normalize(list_images_features, pca_dim=PCA_DIM, scaler=STANDARD_SCALER):
    
    X = np.concatenate(list_images_features, axis=0)  # shape (total_boxes, D)
    X_scaled = (scaler.fit_transform(X)).astype(np.float64)

    print("X_scaled shape:", X_scaled.shape)
    print("Nan ?", np.isnan(X_scaled).any())
    print("Inf ?", np.isinf(X_scaled).any())
    print("Max abs:", np.max(np.abs(X_scaled)))

    pca = PCA(n_components=pca_dim)
    X_scaled_pca = pca.fit_transform(X_scaled)
    
    return X_scaled_pca



# Train a GMM Classification model
def training(X_scaled, gmm_config, output_model_path="output_model.pkl"):
    
    print("GMM training...")
    start_time = time.time()

    gmm_config.fit(X_scaled)
    end_time = time.time()
    total_time = end_time - start_time
    nbr_clusters_found = (gmm_config.weights_ > 1e-3).sum()  
    
    returns = {
        'model': gmm_config,
        'model_path': output_model_path,
        'training_time': total_time,
        'nbr_clusters': nbr_clusters_found,
    }
    
    return returns


def pipeline_training(h5_path, **kwargs):
    """
        This function will perform a full pipeline training without any configuration
        Usefull if you don't know what to do or for a first model
        For advanced configured training, please check the tuto on the github page
        
        input:
            h5_path: Path to the images features to train on
            -- model_path: Path to the where the model should be saved
            -- n_clusters: Number of clusters to find approximately
            -- pca_dim: Number of dimensions the features will be reduct (PCA) (must be the same bteween training and processing)
            -- sacler: The scaler used to normalize data before treatment (default: StandardScaler)
            
        output:
            gmm_model: the gmm model trained
    """
    
    output_model_path = kwargs.get("model_path", "model_classif.joblib")
    output_scaler_path = kwargs.get("scaler_output", "scaler_classif.joblib")
    n_clusters = kwargs.get("nbr_clusters", 4)
    pca_dim = kwargs.get("pca_dim", PCA_DIM)
    scaler = kwargs.get("scaler", STANDARD_SCALER)

    # === Étape 1: Charger tous les box_features ===
    print("Extracting box features...")
    all_box_features, _ = extract_h5(h5_path)

    # === Étape 2: Normalisation ===
    print("Data normalisation...")
    X = np.concatenate(all_box_features, axis=0)
    pipe = Pipeline([
        ("scaler", scaler),
        ("pca", PCA(n_components=pca_dim, svd_solver="auto", random_state=0)),
    ])
    X_scaled = pipe.fit_transform(X)


    # === Étape 3: Configuration du modele GMM a entrainer ===
    gmm = GaussianMixture(
        n_components=n_clusters, 
        covariance_type='full', 
        reg_covar=1e-4, 
        random_state=42
    )


    # === Étape 4: Entrainement du modele ===
    gmm_infos = training(X_scaled, gmm, output_model_path)


    # === Étape 5: Analyse du model ===
    gmm_model = gmm_infos['model']
    ouptut_path = gmm_infos['model_path']
    total_time = gmm_infos['training_time']
    nbr_clusters_found = gmm_infos['nbr_clusters']
    
    
    # === Étape 6: Sauvegarde du modèle et du scaler ===
    print('Saving model and scaler...')
    joblib.dump(gmm_model, output_model_path)
    joblib.dump(pipe, output_scaler_path)


    # === Étape 7: Prints finaux d'infos utiles ===
    print(f"✅ GMM model trained and saved here: {ouptut_path}")
    print(f"✅ GMM scaler trained and saved here: {output_scaler_path}")
    print(f"{nbr_clusters_found} clusters found in training_time {total_time:.3f}s)")
    print("Cluster weights:", gmm_model.weights_)
    print("Means shape:", gmm_model.means_.shape)
    print("Covariances shape:", gmm_model.covariances_.shape)
    
    return gmm_model




# Classifier un batch de features d'images (total_boxes, D)
def classify_batch(features, model, scaler):
    """
        This function will classify each detected cristals from it
        
        input:
            features: numpy array shape (N, D)
            -- model_path: Path to the model to be used to classify
            -- scaler_path: Path to the scaler to be used to normalize data beforehand
            
        output:
            clusters: list(int) List of class_id predicted
    """
    if features.shape[0] == 0:
        return np.array([])
    elif features.ndim == 1:
        features = features.reshape(1, -1)  # devient (1, 1024)
    
    # Normalize data
    X_scaled = scaler.transform(features)
    
    # Predict clusters
    clusters = model.predict(X_scaled)
    
    return clusters




# Enregistrer les clusters dans un csv
# NOTE: Soon to be deprecated as we will less do the features extraction and classification separtely
#       See the colab notebook Detection&Classification
def clusters2csv(image_names, clusters, output_csv):
    """
    Args:
        image_names: list of image names
        all_box_features: list of np.arrays (N_i, D) per image
        clusters_list_of_lists: list of lists of cristal indices per cluster
        output_csv: output csv filepath
    """

    with open(output_csv, 'w', newline='') as f:
        writer = csv.writer(f)
        
        for img_name, cluster in zip(image_names, clusters):
            writer.writerow([img_name])
            n_cristaux = cluster.shape[0]
            for i in range(n_cristaux):                  
                cristal_name = f"cristal{i+1}"             
                writer.writerow([cristal_name, cluster[i]])
                
                
# full pipeline to classify images fetaures from a h5 file
# NOTE: Soon to be deprecated as we will less do the features extraction and classification separtely
#       See the colab notebook Detection&Classification              
def pipeline_process(h5_path, model_path, scaler_path, **kwargs):
    
    output_csv = kwargs.get("output_csv", "results_classif.csv")
    
    print("Openning file...")
    all_box_features, image_names = extract_h5(h5_path)
    
    list_nbr_cristaux = [arr.shape[0] for arr in all_box_features]
    nbr_images = len(image_names)
    
    X = np.concatenate(all_box_features, axis=0)
    
    start_time = time.time() 
    model = joblib.load(model_path)
    scaler = joblib.load(scaler_path)
    clusters = classify_batch(X, model, scaler)
    total_time = time.time() - start_time
    
    # RErepartir les cristaux par images
    count = 0
    list_clusters = []
    for N_i in list_nbr_cristaux:
        clusters_img = clusters[count:count+N_i]
        list_clusters.append(clusters_img)
        count = count+N_i
        
    # Enregistrement des clusters predits au format CSV
    print("Saving predictions...")
    clusters2csv(image_names, list_clusters, output_csv)
    
    # Affichage des statistiques
    print(f"\n✅ Clusters saved successfully: {output_csv}")
    print(f"   Classification summary: \n{nbr_images} images processed in {total_time:.3f}s")
    print(f"Repartition: {proportions(clusters)}") 
    
    image_to_clusters = dict(zip(image_names, list_clusters))
    
    return image_to_clusters
    

# All useful functions to save the cristals sizes, class etc.. in the way FIlip wants
# Processing differenlty for crystal images and times images
class Filip_Saver():
    """
    Args
        all_images_info: Le dictionnaire d'infos de chaque image obtenu avec Detection&Classification
        output_dir: Base folder in wich save the xlsx (Default .)
    """
    
    def __init__(self, all_images_info, root_dir, output_dir='output'):
        self.all_images_info = all_images_info
        self.root_dir = root_dir
        self.output_dir = os.path.basename(root_dir) + "_output/"
        
        path = Path(self.output_dir)
        path.mkdir(parents=True, exist_ok=True)
    
        
    def gestion_crystal_images(self, crystal_folder):
        """
        Gere la creation d un fichier xlsx pour eregistrer les infos de cristaux d'images d etype 'crystal images'
        Create a file for the folder of crystals images in input with:
            1 tab per image 
            
        Args
            crystal_folder: The input folder of crystals images to save
            
        Returns
            None
        """
        print(self.root_dir)
        parent = os.path.basename(os.path.dirname(crystal_folder))
        filename = parent + ".xlsx"
        
        rel_path = os.path.relpath(crystal_folder, self.root_dir)
        reduct_path = os.path.dirname(os.path.dirname(rel_path))
        save_path = os.path.join(self.output_dir, reduct_path)
        os.makedirs(save_path, exist_ok=True)

        xlsx_path = os.path.join(save_path, filename)
        
        wb = Workbook()
        ws = wb.active
        first_page = True
        

        list_files = list(Path(crystal_folder).rglob('*'))
        for image_info in self.all_images_info:
            if any(os.path.basename(image_info["name"]) == p.name for p in list_files):
                
                # Creation de la feuille avec le bon nom
                if first_page:
                    ws.title = os.path.basename(image_info["name"])
                    first_page = False
                else:
                    ws = wb.create_sheet(os.path.basename(image_info["name"]))
                    
                # Enregistrement des infos dans le fichier
                ws.append([image_info["name"], "", "", ""])
                ws.append(["Cristal Id", "size (pixels²)", "size (µm²)", "Class"])
                for crystal_count, row in enumerate(image_info["crystal_info"]):
                    ws.append(row)

                ws[f"A{crystal_count+4}"] = f"AVG"
                ws[f"B{crystal_count+4}"] = f"=AVERAGE(B3:B{crystal_count+3})"
                ws[f"C{crystal_count+4}"] = f"=AVERAGE(C3:C{crystal_count+3})"

        # Enregistrement du fichier
        wb.save(xlsx_path)
        print(f"✅ {xlsx_path} file (Crystal Image) created")
        
        return

    
    def gestion_tab_time(self, ws, time_folder_path):
        
        print('Gestion:', time_folder_path)
        
        arrays = []
        ligne1 = []
        ligne2 = []
        
        # Récupérer toutes les images png/jpg/jpeg
        root_dir = Path(time_folder_path)
        list_files = list(root_dir.rglob("*.png")) + \
                    list(root_dir.rglob("*.jpg")) + \
                    list(root_dir.rglob("*.jpeg"))
        print(list_files)
        for image_info in self.all_images_info:
            if image_info["name"] in list_files:
                ligne1.extend(image_info["name"], ' ', ' ', ' ')
                ligne2.extend("Cristal Id", "size (pixels²)", "size (µm²)", "Class")
                crystal_arr = np.array(image_info["crystal_info"])
                arrays.append(crystal_arr)
        
        
        # Calculer la longueur maximale
        maxN = max(arr.shape[0] for arr in arrays)

        # Créer un tableau vide rempli de ' ' (dtype=object pour pouvoir stocker des strings)
        result = np.full((maxN, 4), ' ', dtype=object)

        # Copier chaque array dans le résultat
        for i, arr in enumerate(arrays):
            N = arr.shape[0]
            result[:N, :] = arr  # remplit les N premières lignes
        
        
        ws.append(ligne1)
        ws.append(ligne2)
        ws.append(result)
        
        
        
        return
    
    
    def gestion_time_images(self, grouped):
        
        for gp, files in grouped.items():
            
            print(f"{gp}:")
            for f in files:
                print(f"  - {f}")
                
            # Creer fichier xlsx avec nom grandparent
            xlsx_path = self.output_dir + gp + ".xlsx"
            wb = Workbook()
            ws = wb.active
            first_page = True
            
            # Creer tabs avec noms parents
            for f in files:
                
                # Creatuion de la feuille avec le bon nom
                parent_name = os.path.basename(os.path.dirname(f))
                if first_page:
                    ws.title = parent_name
                else:
                    ws = wb.create_sheet(parent_name)
                
                # Gestion tab pour chaque time_folder 
                self.gestion_tab_time(ws, f)
                
            print(f"✅ {xlsx_path} file (Time Folder grand-parent) created")
                
        return
    
        
                


def filip_save(all_images_info, root_dir, output_dir="."):
    """
    This function starting from args will properly extract infos from it s path to correclt save it to xlsx according
    to Filip demands
    
    Args:
        all_images_info: list of image_info dicts
            ex: image_info = {"name": img_path, "crystal_info": crystal_info, "image": out.get_image()}
        root_dir: The folder in its original structure to save data from
        output_dir: output_dir: Base folder in wich recretae the structure of root_dir and save all xlsx files
    """
    
    saver = Filip_Saver(all_images_info, root_dir, output_dir)
    
    grouped = defaultdict(list)
    root_dir = Path(root_dir)
    
    for path in root_dir.rglob('*'):
        if not path.is_dir():
            continue
        if path.name == "crystal images":
            saver.gestion_crystal_images(path)
        elif path.name == "time images":
            grandparent = os.path.basename(os.path.dirname(os.path.dirname(path)))
            grouped[grandparent].append(path)
            
    # # Affichage
    # for gp, files in grouped.items():
    #     print(f"{gp}:")
    #     for f in files:
    #         print(f"  - {f}")
    return
    saver.gestion_time_images(grouped)
        

        
    
    