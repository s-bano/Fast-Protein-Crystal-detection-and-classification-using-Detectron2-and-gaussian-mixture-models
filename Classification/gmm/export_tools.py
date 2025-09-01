from pathlib import Path
from openpyxl import Workbook
from collections import defaultdict
from typing import Tuple
from collections import Counter
import matplotlib.pyplot as plt
import numpy as np
import os



def getStats(all_images_infos):
    clusters_all = []
    for image_info in all_images_infos:
        for crystal in image_info["crystal_info"]:
            clusters_all.append(crystal[3])
    count = Counter(clusters_all)
    total = len(clusters_all)
    return {
        k: round((v / total) * 100, 2) 
        for k, v in count.items()
    }


def split_path_at(path: str, folder: str) -> Tuple[Path, Path]:
    """
    Coupe un chemin en deux parties autour d'un dossier cible.
    
    Args:
        path: chemin complet (str ou Path)
        folder: nom du dossier à utiliser comme pivot

    Returns:
        Tuple(before, after):
        - before: Path des dossiers avant le dossier pivot
        - after: Path des dossiers après le dossier pivot
    """
    path = Path(path)
    parts = path.parts
    if folder not in parts:
        raise ValueError(f"Le dossier '{folder}' n'est pas dans le chemin '{path}'")
    
    index = parts.index(folder)
    before = Path(*parts[:index])
    after = Path(*parts[index+1:])
    
    return before, after


# This function concatenate a list of list into a list horizontaly
# list of n (Ni, 4) lists into a (max(Ni), 4n)
def concat_horizontal(arrays):
    # arrays est une liste de listes de lignes (chaque "array" est une liste de listes)
    
    # Trouver le nombre maximal de lignes
    maxN = max(len(arr) for arr in arrays)

    # Normaliser la taille de chaque sous-array en ajoutant des ' ' à la fin
    normalized_arrays = []
    for arr in arrays:
        n = len(arr)
        m = len(arr[0]) if arr else 0  # nombre de colonnes (4 dans ton cas)
        arr_extended = arr + [[' '] * m for _ in range(maxN - n)]
        normalized_arrays.append(arr_extended)

    # Concaténer horizontalement ligne par ligne
    result = []
    for row_idx in range(maxN):
        row = []
        for arr in normalized_arrays:
            row.extend(arr[row_idx])
        result.append(row)

    return result



def image_info_to_arr(image_info, do_avg=False):
    
    array = []
    
    ligne1 = [image_info["name"], ' ', ' ', ' ']
    ligne2 = ["Cristal Id", "size (pixels²)", "size (µm²)", "Class"]
    
    if do_avg:
        # Creation de la ligne des moyennes
        crystal_count = len(image_info["crystal_info"])
        ligne_finale = ["AVG", f"=AVERAGE(B3:B{crystal_count+2})", f"=AVERAGE(C3:C{crystal_count+3})", " "]
    else:
        ligne_finale = [" ", " ", " ", " "]
       
    array = [ligne1, ligne2]  
    
    for row in image_info["crystal_info"]:
        array.append(row) 
    array.append(ligne_finale)
     
    return array
        

# All useful functions to save the cristals sizes, class etc.. in the way FIlip wants
# Processing differenlty for crystal images and times images
class Filip_Saver():
    """
    Args
        all_images_info: Le dictionnaire d'infos de chaque image obtenu avec Detection&Classification
        output_dir: Base folder in wich save the xlsx (Default .)
    """
    
    def __init__(self, all_images_info, root_dir, save_visual):
        self.all_images_info = all_images_info
        self.root_dir = root_dir
        self.output_dir = os.path.basename(root_dir) + "_output/"
        self.crystal_files = 0
        self.times_files = 0
        self.images_created = 0
        self.save_visual = save_visual
        
        path = Path(self.output_dir)
        path.mkdir(parents=True, exist_ok=True)
    

    # Updated version of gestion_crystal_images
    def gestion_crystal_images(self, crystal_folder):
        
        # Creation du fichier excel
        parent = os.path.basename(os.path.dirname(crystal_folder))
        filename = (parent + "_crystal.xlsx")[-31:]
        
        rel_path = os.path.relpath(crystal_folder, self.root_dir)
        reduct_path = os.path.dirname(os.path.dirname(rel_path))
        save_path = os.path.join(self.output_dir, reduct_path)
        os.makedirs(save_path, exist_ok=True)

        xlsx_path = os.path.join(save_path, filename)
        
        wb = Workbook()
        ws = wb.active
        
        # Parcours des images_infos concernes
        arrays = []
        list_files = list(Path(crystal_folder).rglob('*'))
        for image_info in self.all_images_info:
            if any(os.path.basename(image_info["name"]) == p.name for p in list_files):
                
                # Enregistrement de l'image de visualisation
                if self.save_visual:
                    image_path = os.path.join(self.output_dir, *os.path.normpath(image_info["name"]).split(os.sep)[2:])
                    os.makedirs(os.path.dirname(image_path), exist_ok=True)
                    plt.imsave(image_path, image_info["image"])
                    self.images_created += 1
                
                # Extraction des infos et preparation pour ecriture dans fichier Excel
                arr = image_info_to_arr(image_info)
                arrays.append(arr)
                
                
        
        result = concat_horizontal(arrays)
        
        # Ecrire le resultat dans le xlsx
        for row in result:
            ws.append(row)
            
        wb.save(xlsx_path)
        self.crystal_files += 1
        print(f"Crystal images xlsx file created: {xlsx_path}")
        
        return
    
    
    def gestion_tab_time(self, ws, time_folder_path):
        
        arrays = []
        
        # Récupérer toutes les images png/jpg/jpeg
        root_dir = Path(time_folder_path)
        list_files = list(root_dir.rglob("*.png")) + \
                    list(root_dir.rglob("*.jpg")) + \
                    list(root_dir.rglob("*.jpeg"))

        # Pour chaque image creer un array a concatener lui correspondant
        for image_info in self.all_images_info:
            if any(os.path.basename(image_info["name"]) == p.name for p in list_files):
                
                # Enregistrement de l'image de visualisation
                if self.save_visual:
                    image_path = os.path.join(self.output_dir, *os.path.normpath(image_info["name"]).split(os.sep)[2:])
                    os.makedirs(os.path.dirname(image_path), exist_ok=True)
                    plt.imsave(image_path, image_info["image"])
                    self.images_created += 1
                
                # Extraction des infos et preparation pour ecriture dans fichier Excel
                arr = image_info_to_arr(image_info)
                arrays.append(arr)
            
                
        
        result = concat_horizontal(arrays)
        
        # Ecrire le resultat dans le xlsx
        for row in result:
            ws.append(row)
        
        return
    
    
    def gestion_time_images(self, grouped):
        
        for gp, files in grouped.items():
                
            # Creer fichier xlsx avec nom grandparent
            if gp == "crystal_images_filip":
                xlsx_path = os.path.join(self.output_dir, "control bubbles_time.xlsx")
            else:
                filename = (gp + '_time.xlsx')[-30:]
                before, _ = split_path_at(files[0], gp)
                _, after = split_path_at(before, os.path.basename(self.root_dir))
                xlsx_path = os.path.join(self.output_dir, after, filename)
    
            
            wb = Workbook()
            ws = wb.active
            first_page = True
            
            # Creer tabs avec noms parents
            for f in files:
                
                # Creation de la feuille avec le bon nom
                parent_name = os.path.basename(os.path.dirname(f))
                if first_page:
                    ws.title = parent_name[:30]
                    first_page = False
                else:
                    ws = wb.create_sheet(parent_name[:30])
                
                # Gestion tab pour chaque time_folder 
                self.gestion_tab_time(ws, f)
                
            wb.save(xlsx_path)
            self.times_files += 1
            print(f"Time images xlsx file created: {xlsx_path}")
                
        return
    
        
                


def filip_save(all_images_info, root_dir, save_visual=True):
    """
    This function starting from args will properly extract infos from it s path to correclt save it to xlsx according
    to Filip demands
    
    Args:
        all_images_info: list of image_info dicts
            ex: image_info = {"name": img_path, "crystal_info": crystal_info, "image": out.get_image()}
        root_dir: The folder in its original structure to save data from
        save_visual: Boolean to set to True for the visualisation images to be created
    """
    
    saver = Filip_Saver(all_images_info, root_dir, save_visual)
    
    grouped = defaultdict(list)
    root_dir = Path(root_dir)
    crystal_count = 0
    
    for path in root_dir.rglob('*'):
        if not path.is_dir():
            continue
        if path.name == "crystal images":
            saver.gestion_crystal_images(path)
            crystal_count += 1
        elif path.name == "time images":
            grandparent = os.path.basename(os.path.dirname(os.path.dirname(path)))
            grouped[grandparent].append(path)
            
    # # Affichage
    # for gp, files in grouped.items():
    #     print(f"{gp}:")
    #     for f in files:
    #         print(f"  - {f}")
    saver.gestion_time_images(grouped)
    
    # Affichage du recap du nombre de fichiers crees par categories
    print("✅ Done !")
    print(f"Total: {saver.crystal_files + saver.times_files} xlsx files created")
    print(f"\t - {saver.crystal_files} crystal images xlsx files")
    print(f"\t - {saver.times_files} time images xlsxfiles")
    print(f"Total: {saver.images_created} visualisation images saved")

